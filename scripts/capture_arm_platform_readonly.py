"""Read-only technical capture for one ARM platform account.

The script intentionally keeps credentials in memory only. It does not write
cookies, tokens, request bodies, response bodies, HAR files, or screenshots.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from urllib.parse import parse_qsl, urlparse, urlunparse

import openpyxl
from playwright.sync_api import Browser, Page, TimeoutError as PlaywrightTimeoutError
from playwright.sync_api import sync_playwright


SECRET_HINTS = (
    "password",
    "passwd",
    "pass",
    "contrase",
    "clave",
    "token",
    "secret",
    "cookie",
    "session",
    "authorization",
    "auth",
)

BLOCKED_NAV_LABELS = (
    "eliminar",
    "borrar",
    "delete",
    "remove",
    "guardar",
    "save",
    "subir",
    "upload",
    "cargar",
    "enviar",
    "submit",
    "logout",
    "salir",
    "cerrar sesion",
    "cerrar sesión",
    "descargar",
    "download",
)


@dataclass(frozen=True)
class PlatformCredentials:
    label: str
    raw_url: str
    username: str
    password: str
    notes: str


def mask_username(value: str) -> str:
    value = value.strip()
    if not value:
        return ""
    if "@" in value:
        head, tail = value.split("@", 1)
        return f"{head[:1]}***@{tail}"
    return f"{value[:2]}***"


def redact_text(value: Any) -> str:
    text = "" if value is None else str(value)
    text = re.sub(r"[\w.\-+]+@[\w.\-]+\.\w+", "[email]", text)
    text = re.sub(r"\b\d{8}[A-Za-z]\b", "[dni]", text)
    text = re.sub(r"\b[XYZ]\d{7}[A-Za-z]\b", "[nie]", text, flags=re.I)
    text = re.sub(r"\b(?:\+?\d[\d\s().-]{7,}\d)\b", "[phone-or-id]", text)
    for hint in SECRET_HINTS:
        text = re.sub(
            rf"({hint}\s*[:=]\s*)([^\s,;]+)",
            r"\1[redacted]",
            text,
            flags=re.I,
        )
    return " ".join(text.split())[:240]


def normalize_url(raw_url: str) -> str:
    url = raw_url.strip()
    if not re.match(r"^https?://", url, flags=re.I):
        url = f"https://{url}"
    return url


def sanitize_url(raw_url: str) -> str:
    if not raw_url:
        return ""
    parsed = urlparse(raw_url)
    safe_query = []
    for key, _value in parse_qsl(parsed.query, keep_blank_values=True):
        if any(hint in key.lower() for hint in SECRET_HINTS):
            safe_query.append((key, "[redacted]"))
        else:
            safe_query.append((key, "[value]"))
    query = "&".join(f"{key}={value}" for key, value in safe_query)
    path = re.sub(r"/\d{3,}(?=/|$)", "/[id]", parsed.path)
    path = re.sub(r"/[0-9a-f]{16,}(?=/|$)", "/[hash]", path, flags=re.I)
    return urlunparse((parsed.scheme, parsed.netloc, path, "", query, ""))


def safe_filename(value: str) -> str:
    value = value.lower()
    value = re.sub(r"[^a-z0-9]+", "-", value)
    value = value.strip("-")
    return value[:48] or "platform"


def find_local_chromium() -> Path | None:
    root = Path.home() / "AppData" / "Local" / "ms-playwright"
    if not root.exists():
        return None
    candidates = sorted(
        root.glob("chromium-*/chrome-win*/chrome.exe"),
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )
    return candidates[0] if candidates else None


def load_row(excel_path: Path, sheet_name: str, row_number: int) -> PlatformCredentials:
    wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        raise SystemExit(f"Sheet not found: {sheet_name}")
    ws = wb[sheet_name]
    values = [cell.value for cell in ws[row_number]]
    label = str(values[0] or "").strip()
    raw_url = str(values[1] or "").strip()
    username = str(values[2] or "").strip()
    password = str(values[3] or "").strip()
    notes = str(values[4] or "").strip() if len(values) > 4 else ""
    if not label or not raw_url or not username or not password:
        raise SystemExit(
            "Selected row does not contain the expected label/url/username/password columns"
        )
    return PlatformCredentials(label, raw_url, username, password, notes)


def collect_dom_shape(page: Page) -> dict[str, Any]:
    return page.evaluate(
        """() => {
          const clean = (value) => (value || '').toString().replace(/\\s+/g, ' ').trim().slice(0, 160);
          const inputShape = (input) => ({
            tag: input.tagName.toLowerCase(),
            type: input.getAttribute('type') || '',
            name: input.getAttribute('name') || '',
            id: input.getAttribute('id') || '',
            autocomplete: input.getAttribute('autocomplete') || '',
            placeholder: clean(input.getAttribute('placeholder') || ''),
            ariaLabel: clean(input.getAttribute('aria-label') || ''),
            required: Boolean(input.required),
          });
          const forms = Array.from(document.querySelectorAll('form')).slice(0, 8).map((form) => ({
            method: (form.getAttribute('method') || 'get').toLowerCase(),
            action: form.getAttribute('action') || '',
            inputs: Array.from(form.querySelectorAll('input, select, textarea')).slice(0, 40).map(inputShape),
            buttons: Array.from(form.querySelectorAll('button, input[type=submit], input[type=button]')).slice(0, 20).map((button) => ({
              tag: button.tagName.toLowerCase(),
              type: button.getAttribute('type') || '',
              text: clean(button.innerText || button.value || button.getAttribute('aria-label') || ''),
              id: button.getAttribute('id') || '',
              name: button.getAttribute('name') || '',
            })),
          }));
          const navLinks = Array.from(document.querySelectorAll('nav a, aside a, header a, [role=navigation] a, .menu a, .navbar a, .sidebar a'))
            .slice(0, 80)
            .map((a) => ({ text: clean(a.innerText || a.getAttribute('aria-label') || a.title || ''), href: a.href || '' }))
            .filter((a) => a.text || a.href);
          const headings = Array.from(document.querySelectorAll('h1, h2, h3')).slice(0, 20).map((h) => clean(h.innerText));
          const tableHeaders = Array.from(document.querySelectorAll('table')).slice(0, 12).map((table) =>
            Array.from(table.querySelectorAll('thead th, tr:first-child th')).slice(0, 20).map((h) => clean(h.innerText))
          ).filter((headers) => headers.length);
          const buttons = Array.from(document.querySelectorAll('button, a.btn, input[type=submit], input[type=button]')).slice(0, 60).map((button) => clean(button.innerText || button.value || button.getAttribute('aria-label') || '')).filter(Boolean);
          const captchaSignals = Array.from(document.querySelectorAll('iframe, script, div, input')).slice(0, 500).some((el) => {
            const value = `${el.getAttribute('src') || ''} ${el.getAttribute('class') || ''} ${el.getAttribute('id') || ''} ${el.getAttribute('name') || ''}`.toLowerCase();
            return value.includes('captcha') || value.includes('recaptcha') || value.includes('hcaptcha');
          });
          const text = clean(document.body ? document.body.innerText : '');
          const mfaSignals = /\\b(mfa|2fa|otp|verification code|codigo de verificacion|c[oó]digo de verificaci[oó]n|doble factor)\\b/i.test(text);
          return { forms, navLinks, headings, tableHeaders, buttons, captchaSignals, mfaSignals };
        }"""
    )


def empty_dom_shape() -> dict[str, Any]:
    return {
        "forms": [],
        "navLinks": [],
        "headings": [],
        "tableHeaders": [],
        "buttons": [],
        "captchaSignals": False,
        "mfaSignals": False,
    }


def safe_page_title(page: Page) -> str:
    try:
        return redact_text(page.title())
    except Exception:
        return ""


def find_login_fields(page: Page) -> tuple[Any, Any]:
    password = page.locator("input[type='password']").first
    if password.count() == 0:
        return None, None
    user_selectors = [
        "input[type='email']",
        "input[name*='email' i]",
        "input[id*='email' i]",
        "input[name*='user' i]",
        "input[id*='user' i]",
        "input[name*='login' i]",
        "input[id*='login' i]",
        "input[type='text']",
    ]
    for selector in user_selectors:
        locator = page.locator(selector).first
        if locator.count() > 0 and locator.is_visible():
            return locator, password
    return None, password


def click_submit(page: Page) -> None:
    candidates = [
        "button[type='submit']",
        "input[type='submit']",
        "button:has-text('Acceder')",
        "button:has-text('Entrar')",
        "button:has-text('Iniciar')",
        "button:has-text('Login')",
        "button:has-text('Validar')",
        "text=/^(Acceder|Entrar|Iniciar sesi[oó]n|Login|Validar)$/i",
    ]
    for selector in candidates:
        locator = page.locator(selector).first
        if locator.count() > 0 and locator.is_visible():
            locator.click(timeout=10_000)
            return
    page.keyboard.press("Enter")


def should_visit_link(label: str, href: str, base_netloc: str) -> bool:
    label_l = label.lower()
    href_l = href.lower()
    if any(blocked in label_l or blocked in href_l for blocked in BLOCKED_NAV_LABELS):
        return False
    parsed = urlparse(href)
    if parsed.scheme not in {"http", "https"}:
        return False
    if parsed.netloc != base_netloc:
        return False
    if any(ext in parsed.path.lower() for ext in (".pdf", ".zip", ".xls", ".xlsx", ".csv")):
        return False
    return True


def summarize_requests(requests: list[dict[str, Any]], responses: list[dict[str, Any]]) -> dict[str, Any]:
    hosts = Counter()
    paths_by_host: dict[str, Counter[str]] = defaultdict(Counter)
    methods = Counter()
    statuses = Counter()
    content_types = Counter()
    for item in requests:
        parsed = urlparse(item["url"])
        hosts[parsed.netloc] += 1
        paths_by_host[parsed.netloc][parsed.path or "/"] += 1
        methods[item["method"]] += 1
    for item in responses:
        statuses[str(item["status"])] += 1
        if item.get("content_type"):
            content_types[item["content_type"].split(";")[0].strip().lower()] += 1
    return {
        "hosts": hosts.most_common(20),
        "methods": methods.most_common(),
        "statuses": statuses.most_common(),
        "content_types": content_types.most_common(20),
        "top_paths_by_host": {
            host: counter.most_common(25) for host, counter in list(paths_by_host.items())[:20]
        },
    }


def write_markdown(report: dict[str, Any], path: Path) -> None:
    source = report["source"]
    outcome = report["outcome"]
    network = report["network_summary"]
    pages = report["pages"]
    lines = [
        f"# Captura tecnica ARM: {source['platform_label']}",
        "",
        f"- Fecha UTC: {report['captured_at_utc']}",
        f"- Fuente: hoja `{source['sheet']}`, fila `{source['row']}` del Excel local.",
        f"- URL inicial: `{source['initial_url_sanitized']}`",
        f"- Usuario: `{source['username_masked']}`",
        "- Password: presente en Excel, no persistida.",
        f"- Resultado login: `{outcome['login_status']}`",
        f"- URL final: `{outcome['final_url_sanitized']}`",
        "",
        "## Restricciones",
        "",
        "- Captura de solo lectura; no se han ejecutado acciones de escritura.",
        "- No se han guardado cookies, tokens, contrasenas, cuerpos HTTP, HAR ni capturas internas.",
        "- Si se detecta MFA/captcha, la automatizacion debe detenerse y pedir intervencion humana.",
        "",
        "## Login y controles",
        "",
        f"- Titulo inicial: `{redact_text(outcome.get('initial_title', ''))}`",
        f"- Titulo final: `{redact_text(outcome.get('final_title', ''))}`",
        f"- Captcha detectado: `{outcome.get('captcha_detected')}`",
        f"- MFA/OTP detectado: `{outcome.get('mfa_detected')}`",
        f"- Campo usuario localizado: `{outcome.get('username_field_found')}`",
        f"- Campo password localizado: `{outcome.get('password_field_found')}`",
        "",
        "## Dominios y red observada",
        "",
    ]
    for host, count in network["hosts"]:
        lines.append(f"- `{host}`: {count} requests")
    lines.extend(["", "Estados HTTP observados:"])
    for status, count in network["statuses"]:
        lines.append(f"- `{status}`: {count}")
    lines.extend(["", "## Paginas visitadas", ""])
    for page in pages:
        lines.append(f"### {redact_text(page.get('label') or page.get('title') or page['url_sanitized'])}")
        lines.append(f"- URL: `{page['url_sanitized']}`")
        if page.get("title"):
            lines.append(f"- Titulo: `{redact_text(page['title'])}`")
        if page.get("headings"):
            lines.append("- Encabezados: " + ", ".join(f"`{redact_text(h)}`" for h in page["headings"][:8]))
        if page.get("nav_labels"):
            lines.append("- Navegacion: " + ", ".join(f"`{redact_text(n)}`" for n in page["nav_labels"][:20]))
        if page.get("table_headers"):
            lines.append("- Tablas detectadas:")
            for headers in page["table_headers"][:5]:
                lines.append("  - " + ", ".join(f"`{redact_text(h)}`" for h in headers[:12]))
        if page.get("forms"):
            lines.append("- Formularios detectados:")
            for form in page["forms"][:4]:
                types = [field.get("type") or field.get("tag") for field in form.get("inputs", [])]
                lines.append(
                    f"  - method=`{form.get('method')}`, action=`{sanitize_url(form.get('action', ''))}`, fields={types[:12]}"
                )
        lines.append("")
    lines.extend(
        [
            "## Implicaciones para integracion",
            "",
            "- No hay API oficial confirmada por esta captura. Cualquier conector API requiere contrato/documentacion oficial.",
            "- La automatizacion RPA solo podria plantearse si ARM confirma autorizacion escrita, limites de uso y alcance por cuenta.",
            "- Cualquier escritura futura debe funcionar con `dry_run`, `manual_approval_required` y auditoria antes/despues.",
            "- Faltan por confirmar catalogos oficiales de documentos, acciones permitidas, limites, sandbox y politica contractual.",
        ]
    )
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def capture(args: argparse.Namespace) -> dict[str, Any]:
    excel_path = args.excel or next(Path("requisitos").glob("*.xlsx"))
    creds = load_row(excel_path, args.sheet, args.row)
    initial_url = normalize_url(creds.raw_url)
    parsed_initial = urlparse(initial_url)
    out_dir = args.out_dir / f"arm-{safe_filename(creds.label)}-{datetime.now().strftime('%Y%m%d-%H%M%S')}"
    out_dir.mkdir(parents=True, exist_ok=True)

    requests: list[dict[str, Any]] = []
    responses: list[dict[str, Any]] = []
    pages: list[dict[str, Any]] = []
    outcome: dict[str, Any] = {
        "login_status": "not_attempted",
        "initial_title": "",
        "final_title": "",
        "final_url_sanitized": "",
        "captcha_detected": False,
        "mfa_detected": False,
        "username_field_found": False,
        "password_field_found": False,
        "initial_navigation_error": "",
    }

    with sync_playwright() as playwright:
        executable_path = args.executable_path or find_local_chromium()
        launch_kwargs: dict[str, Any] = {"headless": True}
        if executable_path is not None:
            launch_kwargs["executable_path"] = str(executable_path)
        browser: Browser = playwright.chromium.launch(**launch_kwargs)
        context = browser.new_context(
            viewport={"width": 1366, "height": 900},
            locale="es-ES",
            accept_downloads=False,
        )
        page = context.new_page()

        page.on(
            "request",
            lambda request: requests.append(
                {
                    "method": request.method,
                    "url": sanitize_url(request.url),
                    "resource_type": request.resource_type,
                    "is_navigation": request.is_navigation_request(),
                }
            ),
        )
        page.on(
            "response",
            lambda response: responses.append(
                {
                    "url": sanitize_url(response.url),
                    "status": response.status,
                    "content_type": response.headers.get("content-type", ""),
                    "x_frame_options": response.headers.get("x-frame-options", ""),
                    "csp_present": bool(response.headers.get("content-security-policy")),
                }
            ),
        )

        try:
            page.goto(initial_url, wait_until="domcontentloaded", timeout=args.timeout_ms)
            page.wait_for_load_state("networkidle", timeout=10_000)
        except PlaywrightTimeoutError:
            pass
        except Exception as exc:
            outcome["login_status"] = "initial_navigation_failed"
            outcome["initial_navigation_error"] = redact_text(f"{type(exc).__name__}: {exc}")

        outcome["initial_title"] = safe_page_title(page)
        login_shape = empty_dom_shape()
        if outcome["login_status"] != "initial_navigation_failed":
            login_shape = collect_dom_shape(page)
        outcome["captcha_detected"] = bool(login_shape["captchaSignals"])
        outcome["mfa_detected"] = bool(login_shape["mfaSignals"])
        pages.append(
            {
                "label": "login",
                "url_sanitized": sanitize_url(page.url),
                    "title": safe_page_title(page),
                "headings": [redact_text(h) for h in login_shape["headings"]],
                "nav_labels": [redact_text(link["text"]) for link in login_shape["navLinks"]],
                "table_headers": login_shape["tableHeaders"],
                "forms": login_shape["forms"],
            }
        )

        if outcome["login_status"] == "initial_navigation_failed":
            pass
        elif outcome["captcha_detected"] or outcome["mfa_detected"]:
            outcome["login_status"] = "stopped_control_detected_before_login"
        else:
            username_field, password_field = find_login_fields(page)
            outcome["username_field_found"] = username_field is not None
            outcome["password_field_found"] = password_field is not None
            if username_field is None or password_field is None:
                outcome["login_status"] = "login_form_not_found"
            else:
                username_field.fill(creds.username, timeout=10_000)
                password_field.fill(creds.password, timeout=10_000)
                try:
                    with page.expect_navigation(wait_until="domcontentloaded", timeout=15_000):
                        click_submit(page)
                except PlaywrightTimeoutError:
                    try:
                        click_submit(page)
                    except Exception:
                        pass
                try:
                    page.wait_for_load_state("networkidle", timeout=12_000)
                except PlaywrightTimeoutError:
                    pass
                post_shape = collect_dom_shape(page)
                outcome["captcha_detected"] = outcome["captcha_detected"] or bool(post_shape["captchaSignals"])
                outcome["mfa_detected"] = outcome["mfa_detected"] or bool(post_shape["mfaSignals"])
                has_password_after = bool(post_shape["forms"]) and any(
                    field.get("type") == "password"
                    for form in post_shape["forms"]
                    for field in form.get("inputs", [])
                )
                if outcome["captcha_detected"] or outcome["mfa_detected"]:
                    outcome["login_status"] = "stopped_control_detected_after_login"
                elif has_password_after:
                    outcome["login_status"] = "login_not_confirmed_password_form_still_present"
                else:
                    outcome["login_status"] = "login_likely_success"

                pages.append(
                    {
                        "label": "post-login landing",
                        "url_sanitized": sanitize_url(page.url),
                        "title": safe_page_title(page),
                        "headings": [redact_text(h) for h in post_shape["headings"]],
                        "nav_labels": [redact_text(link["text"]) for link in post_shape["navLinks"]],
                        "table_headers": post_shape["tableHeaders"],
                        "forms": post_shape["forms"],
                    }
                )

                if outcome["login_status"] == "login_likely_success":
                    nav_candidates = []
                    for link in post_shape["navLinks"]:
                        text = redact_text(link.get("text"))
                        href = link.get("href", "")
                        if should_visit_link(text, href, urlparse(page.url).netloc):
                            nav_candidates.append((text, href))
                    seen = {sanitize_url(page.url)}
                    for label, href in nav_candidates[: args.max_pages]:
                        safe_href = sanitize_url(href)
                        if safe_href in seen:
                            continue
                        seen.add(safe_href)
                        try:
                            page.goto(href, wait_until="domcontentloaded", timeout=args.timeout_ms)
                            page.wait_for_load_state("networkidle", timeout=8_000)
                            shape = collect_dom_shape(page)
                            pages.append(
                                {
                                    "label": label,
                                    "url_sanitized": sanitize_url(page.url),
                                    "title": safe_page_title(page),
                                    "headings": [redact_text(h) for h in shape["headings"]],
                                    "nav_labels": [redact_text(link["text"]) for link in shape["navLinks"]],
                                    "table_headers": shape["tableHeaders"],
                                    "forms": shape["forms"],
                                }
                            )
                        except Exception as exc:
                            pages.append(
                                {
                                    "label": label,
                                    "url_sanitized": safe_href,
                                    "error": redact_text(f"{type(exc).__name__}: {exc}"),
                                }
                            )

        outcome["final_title"] = safe_page_title(page)
        outcome["final_url_sanitized"] = sanitize_url(page.url)
        context.close()
        browser.close()

    report = {
        "captured_at_utc": datetime.now(timezone.utc).isoformat(),
        "source": {
            "sheet": args.sheet,
            "row": args.row,
            "platform_label": redact_text(creds.label),
            "notes": redact_text(creds.notes),
            "initial_url_sanitized": sanitize_url(initial_url),
            "initial_host": parsed_initial.netloc,
            "username_masked": mask_username(creds.username),
            "password_present": bool(creds.password),
        },
        "outcome": outcome,
        "network_summary": summarize_requests(requests, responses),
        "requests_sample": requests[:120],
        "responses_sample": responses[:120],
        "pages": pages,
    }
    json_path = out_dir / "technical_capture.redacted.json"
    md_path = out_dir / "technical_capture.redacted.md"
    json_path.write_text(json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8")
    write_markdown(report, md_path)
    report["artifact_paths"] = {"json": str(json_path), "markdown": str(md_path)}
    return report


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--excel", type=Path, default=None)
    parser.add_argument("--sheet", default="ARM")
    parser.add_argument("--row", type=int, required=True)
    parser.add_argument("--out-dir", type=Path, default=Path("artifacts") / "platform-captures")
    parser.add_argument("--timeout-ms", type=int, default=30_000)
    parser.add_argument("--max-pages", type=int, default=8)
    parser.add_argument("--executable-path", type=Path, default=None)
    args = parser.parse_args()
    report = capture(args)
    safe = {
        "platform": report["source"]["platform_label"],
        "login_status": report["outcome"]["login_status"],
        "final_url": report["outcome"]["final_url_sanitized"],
        "markdown": report["artifact_paths"]["markdown"],
        "json": report["artifact_paths"]["json"],
    }
    print(json.dumps(safe, indent=2, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    sys.exit(main())
